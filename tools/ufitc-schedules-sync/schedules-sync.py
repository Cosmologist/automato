#!/usr/bin/env python3
"""
schedules-sync — actualize employee schedule Excel file based on working files.

Working files (./var/ufitc-schedules-sync/):
  - schedules.xlsm  — schedule for the previous month (1 employee = 1 row, starts row 3)
  - employees.docx   — current employee list in a Word table (sorted alphabetically)

Usage:
  ./schedules-sync.py employees   — show employee list from employees.docx
  ./schedules-sync.py schedule    — show employee list from schedules.xlsm
  ./schedules-sync.py diff        — show diff between schedule and employee list
  ./schedules-sync.py clean       — remove employees from schedules.xlsm not in employees.docx
  ./schedules-sync.py             — interactive mode selection
"""

import sys
import re
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Error: openpyxl not installed. Run: pip3 install openpyxl")

try:
    import docx
except ImportError:
    sys.exit("Error: python-docx not installed. Run: pip3 install python-docx")


DATA_DIR = Path(__file__).resolve().parent.parent.parent / "var" / "ufitc-schedules-sync"
SCHEDULES_FILE = DATA_DIR / "schedules.xlsm"
EMPLOYEES_FILE = DATA_DIR / "employees.docx"
SHEET_NAME = "Лист2"

GREEN = "\033[32m"
RED = "\033[31m"
YELLOW = "\033[33m"
CYAN = "\033[36m"
NC = "\033[0m"
BOLD = "\033[1m"


def print_summary():
    print("=" * 60)
    print("schedules-sync — actualize employee schedule")
    print("=" * 60)
    print()
    print("Synchronizes schedules.xlsm with the current employee list")
    print(f"from {DATA_DIR}/")
    print()
    print("Modes:")
    print("  employees   — show employee list from employees.docx")
    print("  schedule    — show employee list from schedules.xlsm")
    print("  diff        — show diff between schedule and employee list")
    print("  clean       — remove employees from schedules.xlsm not in employees.docx")
    print()
    print("Usage examples:")
    print(f"  ./{Path(__file__).name} employees")
    print(f"  ./{Path(__file__).name} schedule")
    print(f"  ./{Path(__file__).name} diff")
    print(f"  ./{Path(__file__).name} clean")
    print(f"  ./{Path(__file__).name}              # interactive mode")
    print("=" * 60)
    print()


def normalize_name(name: str) -> str:
    if not name:
        return ""
    name = str(name)
    name = name.replace("\xa0", " ")
    name = name.replace("ё", "е").replace("Ё", "Е")
    name = name.rstrip(",")
    name = re.sub(r"\s+", " ", name).strip()
    return name


def load_employees_docx(path: Path) -> list[tuple[str, str]]:
    doc = docx.Document(str(path))
    table = doc.tables[0]
    employees = []
    for i, row in enumerate(table.rows):
        if i == 0:
            continue
        raw = row.cells[1].text.strip()
        name = normalize_name(raw)
        label = row.cells[2].text.strip()
        if name:
            employees.append((name, label))
    return employees


def cmd_employees():
    if not EMPLOYEES_FILE.exists():
        sys.exit(f"Error: {EMPLOYEES_FILE} not found")

    print(f"Loading employees from {CYAN}{EMPLOYEES_FILE.name}{NC}...")
    employees = load_employees_docx(EMPLOYEES_FILE)
    print(f"  Found {BOLD}{len(employees)}{NC} employees")
    print()
    print(f"{BOLD}{'#':>4}  {'Full Name':<40}  {'Label':<15}{NC}")
    print("-" * 63)
    for i, (name, label) in enumerate(employees, 1):
        print(f"{i:>4}  {name:<40}  {label}")
    print()


def load_schedules_xlsm(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(str(path), read_only=True, keep_vba=True)
    ws = wb[SHEET_NAME]
    employees = []
    for row in ws.iter_rows(min_row=3, max_col=2):
        cell = row[1]
        if cell.value is None or str(cell.value).strip() == "":
            continue
        name = normalize_name(str(cell.value))
        if name:
            employees.append(name)
    wb.close()
    return employees


def cmd_schedule():
    if not SCHEDULES_FILE.exists():
        sys.exit(f"Error: {SCHEDULES_FILE} not found")

    print(f"Loading schedule from {CYAN}{SCHEDULES_FILE.name}{NC}...")
    employees = load_schedules_xlsm(SCHEDULES_FILE)
    print(f"  Found {BOLD}{len(employees)}{NC} employees")
    print()
    print(f"{BOLD}{'#':>4}  {'Full Name':<40}{NC}")
    print("-" * 46)
    for i, name in enumerate(employees, 1):
        print(f"{i:>4}  {name}")
    print()


def build_diff_pairs(schedules: list[str], employees: list[tuple[str, str]]) -> list[tuple[str | None, str | None, str]]:
    employees_names = [name for name, _ in employees]
    employees_labels = {name: label for name, label in employees}

    schedules_sorted = sorted(schedules, key=str.lower)
    employees_sorted = sorted(employees_names, key=str.lower)
    employees_set = set(employees_sorted)
    schedules_set = set(schedules_sorted)

    pairs: list[tuple[str | None, str | None, str]] = []

    for name in schedules_sorted:
        if name in employees_set:
            pairs.append((name, name, employees_labels.get(name, "")))
        else:
            pairs.append((name, None, ""))

    for name in employees_sorted:
        if name not in schedules_set:
            inserted = False
            for i, pair in enumerate(pairs):
                pair_name = pair[0] if pair[0] is not None else pair[1]
                if pair_name.lower() > name.lower():
                    pairs.insert(i, (None, name, employees_labels.get(name, "")))
                    inserted = True
                    break
            if not inserted:
                pairs.append((None, name, employees_labels.get(name, "")))

    return pairs


def cmd_diff():
    if not SCHEDULES_FILE.exists():
        sys.exit(f"Error: {SCHEDULES_FILE} not found")
    if not EMPLOYEES_FILE.exists():
        sys.exit(f"Error: {EMPLOYEES_FILE} not found")

    print(f"Loading schedule from {CYAN}{SCHEDULES_FILE.name}{NC}...")
    schedules = load_schedules_xlsm(SCHEDULES_FILE)
    print(f"  Found {BOLD}{len(schedules)}{NC} employees")

    print(f"Loading employees from {CYAN}{EMPLOYEES_FILE.name}{NC}...")
    employees = load_employees_docx(EMPLOYEES_FILE)
    print(f"  Found {BOLD}{len(employees)}{NC} employees")

    pairs = build_diff_pairs(schedules, employees)

    matched = sum(1 for s, e, _ in pairs if s is not None and e is not None)
    only_schedule = sum(1 for s, e, _ in pairs if s is not None and e is None)
    only_employees = sum(1 for s, e, _ in pairs if s is None and e is not None)

    print()
    print(f"  {GREEN}Matched:{NC} {matched}")
    print(f"  {RED}Only in schedule:{NC} {only_schedule}")
    print(f"  {YELLOW}Only in employees:{NC} {only_employees}")
    print()
    print(f"{BOLD}{'#':>4}  {'schedule.xlsm':<40}  {'employees.docx':<40}  {'Label':<15}{NC}")
    print("-" * 105)
    for i, (s, e, label) in enumerate(pairs, 1):
        if s is not None and e is not None:
            color = GREEN
        elif s is not None:
            color = RED
        else:
            color = YELLOW
        s_display = s if s else ""
        e_display = e if e else ""
        print(f"{i:>4}  {color}{s_display:<40}{NC}  {color}{e_display:<40}{NC}  {label}")
    print()


def load_schedules_xlsm_with_rows(path: Path) -> list[tuple[int, str]]:
    wb = openpyxl.load_workbook(str(path), read_only=True, keep_vba=True)
    ws = wb[SHEET_NAME]
    employees = []
    for row in ws.iter_rows(min_row=3, max_col=2):
        cell = row[1]
        if cell.value is None or str(cell.value).strip() == "":
            continue
        name = normalize_name(str(cell.value))
        if name:
            employees.append((cell.row, name))
    wb.close()
    return employees


def cmd_clean():
    if not SCHEDULES_FILE.exists():
        sys.exit(f"Error: {SCHEDULES_FILE} not found")
    if not EMPLOYEES_FILE.exists():
        sys.exit(f"Error: {EMPLOYEES_FILE} not found")

    print(f"Loading schedule from {CYAN}{SCHEDULES_FILE.name}{NC}...")
    schedules_with_rows = load_schedules_xlsm_with_rows(SCHEDULES_FILE)
    print(f"  Found {BOLD}{len(schedules_with_rows)}{NC} employees")

    print(f"Loading employees from {CYAN}{EMPLOYEES_FILE.name}{NC}...")
    employees = load_employees_docx(EMPLOYEES_FILE)
    print(f"  Found {BOLD}{len(employees)}{NC} employees")

    employees_names = [name for name, _ in employees]
    employees_set = set(employees_names)

    to_remove = []
    for row_num, name in schedules_with_rows:
        if name not in employees_set:
            to_remove.append((row_num, name))

    if not to_remove:
        print()
        print("Schedule is already clean. No employees to remove.")
        return

    print()
    print(f"{RED}Employees to remove from schedule ({len(to_remove)}):{NC}")
    for row_num, name in to_remove:
        print(f"  {RED}-{NC} {SCHEDULES_FILE.name}:B{row_num} → {name}")

    print()
    confirm = input(f"Remove {len(to_remove)} employees? [y/N]: ").strip().lower()
    if confirm != "y":
        print("Cancelled.")
        return

    wb = openpyxl.load_workbook(str(SCHEDULES_FILE), keep_vba=True)
    ws = wb[SHEET_NAME]

    rows_to_delete = sorted([row_num for row_num, _ in to_remove])

    merges_to_unmerge = []
    merges_to_shift = []

    for merge in list(ws.merged_cells.ranges):
        overlaps = any(r >= merge.min_row and r <= merge.max_row for r in rows_to_delete)
        if overlaps:
            merges_to_unmerge.append(str(merge))
        else:
            rows_before = sum(1 for r in rows_to_delete if r < merge.min_row)
            if rows_before > 0:
                merges_to_shift.append((str(merge), merge, rows_before))

    for m in merges_to_unmerge:
        ws.unmerge_cells(m)

    for orig_str, merge, rows_before in merges_to_shift:
        ws.unmerge_cells(orig_str)
        new_min_row = merge.min_row - rows_before
        new_max_row = merge.max_row - rows_before
        new_range = f'{get_column_letter(merge.min_col)}{new_min_row}:{get_column_letter(merge.max_col)}{new_max_row}'
        ws.merge_cells(new_range)

    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num, 1)

    wb.save(str(SCHEDULES_FILE))
    print()
    print(f"Removed {len(to_remove)} employees from {SCHEDULES_FILE.name}")
    print("Done.")


def interactive_mode() -> str:
    modes = [
        ("employees", "Show employee list from employees.docx"),
        ("schedule", "Show employee list from schedules.xlsm"),
        ("diff", "Show diff between schedule and employee list"),
        ("clean", "Remove employees from schedules.xlsm not in employees.docx"),
    ]
    print("Available modes:")
    for i, (name, desc) in enumerate(modes, 1):
        print(f"  {i}. {name:10s} — {desc}")
    print()
    while True:
        try:
            choice = input("Enter mode number (1): ").strip()
        except (EOFError, KeyboardInterrupt):
            sys.exit(0)
        if choice == "" or choice == "1":
            return "employees"
        if choice == "2":
            return "schedule"
        if choice == "3":
            return "diff"
        if choice == "4":
            return "clean"
        print("Invalid choice. Enter 1, 2, 3 or 4.")


def main():
    print_summary()

    parser = argparse.ArgumentParser(description="Actualize employee schedule")
    parser.add_argument("mode", nargs="?", help="Operation mode")
    args = parser.parse_args()

    mode = args.mode
    if mode is None:
        mode = interactive_mode()

    if mode == "employees":
        cmd_employees()
    elif mode == "schedule":
        cmd_schedule()
    elif mode == "diff":
        cmd_diff()
    elif mode == "clean":
        cmd_clean()
    else:
        sys.exit(f"Unknown mode: {mode}")


if __name__ == "__main__":
    main()
